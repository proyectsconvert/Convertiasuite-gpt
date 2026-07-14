import os
import logging
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from app.core.files_config import BRAND_CONFIG
from app.domain.entities.document_content import DocumentContent
from app.domain.interfaces.document_builder import IDocumentBuilder
from app.services.document_generation.template_engine import TemplateEngine

logger = logging.getLogger(__name__)


class ExcelBuilder(IDocumentBuilder):

    def __init__(self, engine: TemplateEngine):
        self._engine = engine

    @property
    def output_format(self) -> str:
        return "xlsx"

    def build(self, content: DocumentContent) -> bytes:
        brand = content.brand or "convertia"
        excel_cfg = BRAND_CONFIG[brand].get("excel", {})

        # Enriquecer la configuración con el logo y el nombre de la marca
        excel_cfg = dict(excel_cfg)
        excel_cfg["logo_path"] = BRAND_CONFIG[brand]["logos"].get("main")
        excel_cfg["brand_name"] = BRAND_CONFIG[brand].get("nametag", "Convertia")

        try:
            ctx = self._engine.build_excel_context(content)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Eliminar hoja por defecto

            for sheet_ctx in ctx["sheets"]:
                ws = wb.create_sheet(title=sheet_ctx["name"][:31])
                self._write_sheet(ws, sheet_ctx, excel_cfg)

            buffer = BytesIO()
            wb.save(buffer)
            logger.info(f"XLSX generado exitosamente con branding: '{content.title}'")
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Error generando XLSX '{content.title}': {e}")
            raise RuntimeError(f"Error al generar el archivo Excel: {e}") from e

    def _write_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        ctx: dict,
        cfg: dict,
    ) -> None:
        headers = ctx.get("headers", [])
        rows = ctx.get("rows", [])
        summary = ctx.get("summary")

        logo_path = cfg.get("logo_path")
        if logo_path and os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                # Escalar la imagen de forma que su alto sea de 30px manteniendo el aspect ratio
                aspect_ratio = img.width / img.height if img.height else 1
                img.height = 30
                img.width = int(30 * aspect_ratio)
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 36
            except Exception as e:
                logger.warning(f"No se pudo insertar el logo en Excel: {e}")

        ws.row_dimensions[2].height = 15

        sheet_title = ctx.get("name", "Datos")
        title_col = 3 if (logo_path and os.path.exists(logo_path)) else 1
        title_cell = ws.cell(row=1, column=title_col, value=sheet_title.upper())
        title_cell.font = Font(
            name=cfg.get("font_name", "Calibri"),
            bold=True,
            color=cfg.get(
                "header_fill", "011E23"
            ),  # Usar el color primario corporativo
            size=14,
        )
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        header_fill = PatternFill("solid", fgColor=cfg.get("header_fill", "011E23"))
        header_font = Font(
            name=cfg.get("font_name", "Calibri"),
            bold=True,
            color=cfg.get("header_text", "FFFFFF"),
            size=10,
        )
        zebra_fill = PatternFill("solid", fgColor=cfg.get("zebra_fill", "F5F5F5"))
        border_color = cfg.get("border_color", "D9D9D9")
        body_font = Font(name=cfg.get("font_name", "Calibri"), size=9)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color=border_color),
            right=Side(style="thin", color=border_color),
            top=Side(style="thin", color=border_color),
            bottom=Side(style="thin", color=border_color),
        )

        start_row = 3
        ws.row_dimensions[start_row].height = 24  # Alto de la fila de cabecera

        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c_idx, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for r_idx, row in enumerate(rows, start=start_row + 1):
            ws.row_dimensions[r_idx].height = 18  # Alto estándar de fila de datos
            is_zebra = r_idx % 2 == 0
            for c_idx, value in enumerate(row, start=1):
                if c_idx > len(headers):
                    break
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = body_font
                cell.alignment = left_align
                cell.border = thin_border
                if is_zebra:
                    cell.fill = zebra_fill

        # Autoajuste de Ancho de Columnas
        for c_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(c_idx)
            max_length = len(str(header))
            for row in rows:
                if c_idx - 1 < len(row):
                    max_length = max(max_length, len(str(row[c_idx - 1])))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        if summary:
            summary_row = start_row + len(rows) + 2
            ws.row_dimensions[summary_row].height = 20
            ws.cell(row=summary_row, column=1, value="RESUMEN").font = Font(
                bold=True, color="FFFFFF", name=cfg.get("font_name", "Calibri"), size=10
            )
            ws.cell(row=summary_row, column=1).fill = PatternFill(
                "solid", fgColor=cfg.get("header_fill", "011E23")
            )
            ws.cell(row=summary_row, column=1).alignment = center_align
            ws.cell(row=summary_row, column=1).border = thin_border

            col = 2
            for key, val in summary.items():
                cell_key = ws.cell(row=summary_row, column=col, value=str(key))
                cell_key.font = Font(
                    name=cfg.get("font_name", "Calibri"), size=9, bold=True
                )
                cell_key.alignment = left_align
                cell_key.border = thin_border
                cell_key.fill = zebra_fill

                cell_val = ws.cell(row=summary_row, column=col + 1, value=str(val))
                cell_val.font = body_font
                cell_val.alignment = left_align
                cell_val.border = thin_border

                col += 2

        ws.freeze_panes = "A4"
