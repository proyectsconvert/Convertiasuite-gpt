"""
Excel document processor using openpyxl.
Extracts text, tables, and metadata from XLSX/XLS files.
"""

import io
import logging

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from app.domain.entities.document import (
    DocumentType,
    ParsedContent,
    Section,
    Table,
)
from app.domain.interfaces.document_processor import IDocumentProcessor

logger = logging.getLogger(__name__)


class ExcelProcessor(IDocumentProcessor):
    """
    Processes Excel files using openpyxl.
    Extracts tables from each sheet and preserves structure.
    """

    def __init__(self):
        if load_workbook is None:
            raise ImportError("openpyxl is required for Excel processing. Install with: pip install openpyxl")

    @property
    def supported_type(self) -> DocumentType:
        return DocumentType.EXCEL

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".xlsm"]

    async def parse(
        self,
        file_content: bytes,
        filename: str,
    ) -> ParsedContent:
        """
        Parse Excel file and extract all sheet data.
        Each sheet becomes a section with embedded tables.
        """
        try:
            excel_file = io.BytesIO(file_content)
            workbook = load_workbook(excel_file, data_only=True)

            full_text = ""
            sections: list[Section] = []
            tables: list[Table] = []

            # Process each sheet
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                section_content = f"Sheet: {sheet_name}\n"
                
                # Extract all data
                rows_data = []
                headers = None

                for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if not row or all(cell is None for cell in row):
                        continue

                    # Convert None values to empty strings
                    row_data = [str(cell) if cell is not None else "" for cell in row]

                    if row_idx == 0:
                        headers = row_data
                    else:
                        rows_data.append(row_data)

                    # Add to text
                    section_content += " | ".join(row_data) + "\n"

                # Create table if headers found
                if headers and rows_data:
                    table = Table(
                        headers=headers,
                        rows=rows_data,
                        name=sheet_name,
                        metadata={"sheet_index": sheet_idx, "sheet_name": sheet_name}
                    )
                    tables.append(table)

                # Create section for sheet
                if section_content.strip():
                    section = Section(
                        title=sheet_name,
                        content=section_content.strip(),
                        level=1,
                        metadata={"type": "sheet", "sheet_index": sheet_idx}
                    )
                    sections.append(section)
                    full_text += f"## {sheet_name}\n{section_content}\n\n"

            metadata = {
                "sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "tables_extracted": len(tables),
            }

            return ParsedContent(
                text=full_text.strip(),
                sections=sections,
                tables=tables,
                images=[],
                metadata=metadata,
            )

        except Exception as e:
            logger.error(f"Excel parsing error for {filename}: {str(e)}")
            raise ValueError(f"Failed to parse Excel file: {str(e)}")

    def get_metadata(self, file_content: bytes) -> dict:
        """Extract Excel metadata without full parsing."""
        try:
            excel_file = io.BytesIO(file_content)
            workbook = load_workbook(excel_file, data_only=True)

            return {
                "is_valid": True,
                "sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
            }

        except Exception as e:
            logger.error(f"Excel metadata extraction error: {str(e)}")
            return {"is_valid": False, "error": str(e)}
